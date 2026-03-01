from functools import partial

from PySide6.QtCore import Signal, QThreadPool
from PySide6.QtGui import QStandardItemModel, QStandardItem

from app.ui.customSortingMenuWidget import CustomSortingMenuWidget
from app.ui.widgets.ui_customPaymentsWidget import *
from app.models.tableModel import CustomTableViewWithMultiSelection
from app.models.sortingModel import CaseInsensitiveProxyModel, FilterableHeaderView
from app.ui.customCalendarWidget import CustomCalendarDialog
from app.services.workerServices import WorkerServices as WoS
from app.utils.utils import Utils, Worker


class CustomPaymentsWidget(QWidget, Ui_customPaymentsWidget):
    logoutSignal = Signal(bool)

    def __init__(self, mainWindow, user, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.threadPool = QThreadPool()
        self.mainWindow = mainWindow
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.usernameLabel.setText(user)
        self.user = user

        self.paymentsTableView = CustomTableViewWithMultiSelection()
        self.paymentsTableHolder.layout().addWidget(self.paymentsTableView)
        self.tablePaymentsModel = QStandardItemModel()
        self.tablePaymentsNames = ['ID', '№', 'Име', 'Длъжност', 'Цех', 'Зар. дни', 'Прис. Вр.', 'В Поч. дни',
                                   'В Поч. дни Лв.', 'В Поч. дни €', 'В Празници', 'В Празници Лв.', 'В Празници €',
                                   'Бр.', 'Вр. Опер.', 'Почас.', 'Извънр.', 'Извънр. Лв.', 'Извънр. €', 'Нощен',
                                   'Нощен в Лв.', 'Нощен в €', 'Ефек. в %', 'Извънр. Тотал', 'Извънр. Тотал Лв.',
                                   'Извънр. Тотал €', 'Зар. Лв.', 'Зар. €', 'Тотал Лв.', 'Тотал €']

        # ['ID'0, '№'1, 'Име'2, 'Длъжност'3, 'Цех'4, 'Зар. дни'5, 'Прис. вр.'6, 'В Поч. дни'7,
        #  'В Поч. дни Лв.'8, 'В Поч. дни €'9, 'В Празници'10, 'В Празници Лв.'11, 'В Празници €'12,
        #  'Бр.'13, 'Вр. Опер.'14, 'Почас.'15, 'Извънр.'16, 'Извънр. Лв.'17, 'Извънр. €'18, 'Нощен'19,
        #  'Нощен в Лв.'20, 'Нощен в €'21, 'Ефек. в %'22, 'Извънр. Тотал'23, 'Извънр. Тотал Лв.'24,
        #  'Извънр. Тотал €'25, 'Зар. Лв.'26, 'Зар. €'27, 'Тотал Лв.'28, 'Тотал €'29]
        for i, tableHeaderName in enumerate(self.tablePaymentsNames):
            self.tablePaymentsModel.setHorizontalHeaderItem(i, QStandardItem(tableHeaderName))
            self.tablePaymentsModel.horizontalHeaderItem(i).setTextAlignment(Qt.AlignmentFlag.AlignLeft)
            self.tablePaymentsModel.horizontalHeaderItem(i).setTextAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.proxyModelPaymentsTable = CaseInsensitiveProxyModel(numericColumns=[0, 1, 5, 6, 7, 8, 9, 10, 11, 12,
                                                                                 13, 14, 15, 16, 17, 18, 19, 20,
                                                                                 21, 22, 23, 24, 25, 26, 27, 28, 29],
                                                                 parent=self)
        self.setProxyModel(self.proxyModelPaymentsTable, self.tablePaymentsModel, self.paymentsTableView)
        # self.timePapersForDayTableView.setModel(self.tablePaymentsModel)
        self.paymentsTableView.horizontalHeader().setDefaultSectionSize(80)
        self.paymentsTableView.horizontalHeader().setStretchLastSection(True)
        self.setColumnWidths()

        self.setCurrentPayments()

        self.paymentsTableView.selectedRows.connect(self.showPaymentDetails)
        self.paymentsTableView.clearCurrentSelection.connect(self.resetSelectedInfo)

        # self.fromDateEdit.setDate(QDate.currentDate().addDays(-7))
        # self.toDateEdit.setDate(QDate.currentDate())

        self.checkBoxFiltering = {}
        self.initialCheckBoxes = {}

        self.levaCheckBox.stateChanged.connect(self.onLevaStateChanged)
        self.weekendDaysCheckBox.stateChanged.connect(self.onWeekendStateChanged)
        self.holidaysCheckBox.stateChanged.connect(self.onHolidaysStateChanged)
        self.hourlyCheckBox.stateChanged.connect(self.onHourlyStateChanged)
        self.overtimeCheckBox.stateChanged.connect(self.onOvertimeStateChanged)
        self.nightTimeCheckBox.stateChanged.connect(self.onNightTimeStateChanged)
        self.positionCheckBox.stateChanged.connect(self.onPositionStateChanged)
        self.placeCheckBox.stateChanged.connect(self.onPlaceStateChanged)
        self.selectAllCheckBox.stateChanged.connect(self.onSelectAllStateChanged)

        self.exportToExcelBtn.clicked.connect(self.exportToExcel)

        currentMonth = QDate.currentDate().month()
        daysInCurrentMonth = QDate.daysInMonth(QDate.currentDate())
        startDate = QDate(QDate.currentDate().year(), currentMonth, 1)
        endDate = QDate(QDate.currentDate().year(), currentMonth, daysInCurrentMonth)
        self.forMonthComboBox.setCurrentIndex(currentMonth - 1)
        self.fromDateEdit.setDate(startDate)
        self.toDateEdit.setDate(endDate)

        self.refreshPaymentsTable(startDate, endDate)
        self.setInitialColumns()

        self.forMonthComboBox.currentIndexChanged.connect(self.refreshPaymentsTableForMonth)

        self.fromCalendarBtn.clicked.connect(self.showCalendarDialog)
        self.toCalendarBtn.clicked.connect(self.showCalendarDialog)

        self.fromDateEdit.dateChanged.connect(self.checkFromDate)
        self.toDateEdit.dateChanged.connect(self.checkToDate)

        self.searchBtn.clicked.connect(self.refreshPaymentsTable)
        self.paymentsTableView.doubleClicked.connect(self.openPaymentDetails)

        self.closeBtn.clicked.connect(self.close)
        self.logoutBtn.clicked.connect(self.logout)

    # def setPayPerMinComboBox(self):
    #     paymentsPerMin = WoS.getPaymentsForMin()
    #     self.payPerMinComboBox.clear()

    def refreshPaymentsTableForMonth(self):
        month = self.forMonthComboBox.currentIndex() + 1
        startDate = QDate(QDate.currentDate().year(), month, 1)
        endDate = QDate(QDate.currentDate().year(), month, QDate.daysInMonth(startDate))
        self.fromDateEdit.setDate(startDate)
        self.toDateEdit.setDate(endDate)
        self.refreshPaymentsTable(startDate, endDate)

    def setCurrentPayments(self):
        currentPayments = WoS.getCurrentPayments()
        if currentPayments:
            self.levaPerMinDayLabel.setText(currentPayments['dayInLeva'])
            self.euroPerMinDayLabel.setText(currentPayments['dayInEuro'])
            self.levaPerMinNightLabel.setText(currentPayments['nightInLeva'])
            self.euroPerMinNightLabel.setText(currentPayments['nightInEuro'])
        else:
            self.levaPerMinDayLabel.setText("--")
            self.euroPerMinDayLabel.setText("--")
            self.levaPerMinNightLabel.setText("--")
            self.euroPerMinNightLabel.setText("--")

    def exportToExcel(self):
        """Export the current table view data to Excel file"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import os
        from datetime import datetime
        from PySide6.QtWidgets import QFileDialog

        # Get save file location from user
        fileName, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.join(os.path.expanduser("~"), "Documents",
                         f"Payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            "Excel Files (*.xlsx)"
        )

        if not fileName:
            return  # User canceled

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payments Report"

        # Add report title and date range
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"Payments Report ({self.fromDateEdit.date().toString('dd.MM.yyyy')} - {self.toDateEdit.date().toString('dd.MM.yyyy')})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Add headers
        header_fill = PatternFill(start_color="DFDFDF", end_color="DFDFDF", fill_type="solid")

        columnHeaders = self.tablePaymentsNames.copy()
        if not self.positionCheckBox.isChecked():
            columnHeaders.remove('Длъжност')

        if not self.placeCheckBox.isChecked():
            columnHeaders.remove('Цех')

        if not self.holidaysCheckBox.isChecked():
            columnHeaders.remove('В Празници')
            columnHeaders.remove('В Празници €')
            columnHeaders.remove('В Празници Лв.')

        if self.holidaysCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('В Празници Лв.')

        if not self.weekendDaysCheckBox.isChecked():
            columnHeaders.remove('В Поч. дни')
            columnHeaders.remove('В Поч. дни Лв.')
            columnHeaders.remove('В Поч. дни €')

        if self.weekendDaysCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('В Поч. дни Лв.')

        if not self.overtimeCheckBox.isChecked():
            columnHeaders.remove('Извънр.')
            columnHeaders.remove('Извънр. Лв.')
            columnHeaders.remove('Извънр. €')

        if self.overtimeCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('Извънр. Лв.')

        if not self.nightTimeCheckBox.isChecked():
            columnHeaders.remove('Нощен')
            columnHeaders.remove('Нощен в Лв.')
            columnHeaders.remove('Нощен в €')

        if self.nightTimeCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('Нощен в Лв.')

        if not self.hourlyCheckBox.isChecked():
            columnHeaders.remove('Почас.')

        if not self.levaCheckBox.isChecked():
            columnHeaders.remove('Извънр. Тотал Лв.')
            columnHeaders.remove('Зар. Лв.')
            columnHeaders.remove('Тотал Лв.')

        for col_idx, header_text in enumerate(columnHeaders):
            col_letter = get_column_letter(col_idx + 1)
            cell = ws[f"{col_letter}3"]
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        # self.excelTableNames = self.tablePaymentsNames

        # Add data rows
        row_idx = 4  # Start from row 4 (after headers)
        for row in range(self.proxyModelPaymentsTable.rowCount()):
            viewColumns = 0
            for col in range(self.proxyModelPaymentsTable.columnCount()):
                # Skip hidden columns
                if self.paymentsTableView.isColumnHidden(col):
                    continue

                # Get data from the proxy model
                index = self.proxyModelPaymentsTable.index(row, col)
                value = self.proxyModelPaymentsTable.data(index, Qt.ItemDataRole.DisplayRole)

                # ['ID'0, '№'1, 'Име'2, 'Длъжност'3, 'Цех'4, 'Зар. дни'5, 'Прис. вр.'6, 'В Поч. дни'7,
                #  'В Поч. дни Лв.'8, 'В Поч. дни €'9, 'В Празници'10, 'В Празници Лв.'11, 'В Празници €'12,
                #  'Бр.'13, 'Вр. Опер.'14, 'Почас.'15, 'Извънр.'16, 'Извънр. Лв.'17, 'Извънр. €'18, 'Нощен'19,
                #  'Нощен в Лв.'20, 'Нощен в €'21, 'Ефек. в %'22, 'Извънр. Тотал'23, 'Извънр. Тотал Лв.'24,
                #  'Извънр. Тотал €'25, 'Зар. Лв.'26, 'Зар. €'27, 'Тотал Лв.'28, 'Тотал €'29]

                # Convert to appropriate type
                if col in [0, 1, 5, 13, 22]:  # Integer columns
                    try:
                        value = int(value) if value else 0
                    except ValueError:
                        pass
                elif col in [6, 7, 8, 9, 10, 11, 12, 14, 15, 16, 17,
                             18, 19, 20, 21, 23, 24, 25, 26, 27, 28, 29]:  # Float columns
                    try:
                        value = float(value) if value else 0.0
                    except ValueError:
                        pass

                # Write to cell
                ws.cell(row=row_idx, column=viewColumns + 1, value=value)
                viewColumns += 1
                # print(f"View column: {viewColumns}, Value: {value}")

            row_idx += 1

        # Add summary at the bottom
        summary_row = row_idx + 1
        ws.cell(row=summary_row, column=1, value="Total:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=self.totalViewRows.text()).font = Font(bold=True)

        # Add SUM formulas for Leva (column 13) and Euro (column 14)
        leva_col = get_column_letter(len(columnHeaders) - 1)  # Column M in Excel (index 12 + 1)
        euro_col = get_column_letter(len(columnHeaders))  # Column N in Excel (index 13 + 1)

        # Create SUM formulas that reference the data range
        leva_formula = f"=SUM({leva_col}4:{leva_col}{row_idx})"
        euro_formula = f"=SUM({euro_col}4:{euro_col}{row_idx})"

        # Add the formulas to the summary row
        leva_cell = ws.cell(row=summary_row, column=len(columnHeaders) - 1, value=leva_formula)
        euro_cell = ws.cell(row=summary_row, column=len(columnHeaders), value=euro_formula)

        # Apply bold formatting to the summary cells
        leva_cell.font = Font(bold=True)
        euro_cell.font = Font(bold=True)

        # Save the workbook
        try:
            wb.save(fileName)
            from app.ui.messagesManager import MessageManager as MM
            MM.showOnWidget(self, f"Успешен запис на {fileName}", "success")
        except Exception as e:
            from app.ui.messagesManager import MessageManager as MM
            MM.showOnWidget(self, f"Грешка при записване: {str(e)}", "error")

    def setInitialColumns(self):
        for i in range(len(self.tablePaymentsNames)):
            if i in [0, 1, 2, 5, 6, 13, 14, 22, 23, 25, 27, 29]:
                pass
            else:
                self.paymentsTableView.setColumnHidden(i, True)

    def checkAllstate(self):
        if (self.hourlyCheckBox.isChecked() and
                self.overtimeCheckBox.isChecked() and
                self.nightTimeCheckBox.isChecked() and
                self.weekendDaysCheckBox.isChecked() and
                self.holidaysCheckBox.isChecked() and
                self.levaCheckBox.isChecked() and
                self.positionCheckBox.isChecked() and
                self.placeCheckBox.isChecked()):
            self.selectAllCheckBox.setChecked(True)
        else:
            self.selectAllCheckBox.blockSignals(True)
            self.selectAllCheckBox.setChecked(False)
            self.selectAllCheckBox.blockSignals(False)

    def onSelectAllStateChanged(self, state):
        self.hourlyCheckBox.setChecked(state)
        self.overtimeCheckBox.setChecked(state)
        self.nightTimeCheckBox.setChecked(state)
        self.weekendDaysCheckBox.setChecked(state)
        self.holidaysCheckBox.setChecked(state)
        self.levaCheckBox.setChecked(state)
        self.positionCheckBox.setChecked(state)
        self.placeCheckBox.setChecked(state)

    def onLevaStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(24, not state)
        self.paymentsTableView.setColumnHidden(26, not state)
        self.paymentsTableView.setColumnHidden(28, not state)

        if self.overtimeCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(17, not state)

        if self.nightTimeCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(20, not state)

        if self.holidaysCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(11, not state)

        if self.weekendDaysCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(8, not state)

        self.checkAllstate()

    def onHourlyStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(15, not state)
        self.checkAllstate()

    def onOvertimeStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(16, not state)
        self.paymentsTableView.setColumnHidden(18, not state)
        if self.levaCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(17, False)

        if not self.overtimeCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(17, True)
        self.checkAllstate()

    def onNightTimeStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(19, not state)
        self.paymentsTableView.setColumnHidden(21, not state)

        if self.levaCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(20, False)

        if not self.nightTimeCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(20, True)

        self.checkAllstate()

    def onWeekendStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(7, not state)
        self.paymentsTableView.setColumnHidden(9, not state)

        if self.levaCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(8, False)

        if not self.weekendDaysCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(8, True)

        self.checkAllstate()

    def onHolidaysStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(10, not state)
        self.paymentsTableView.setColumnHidden(12, not state)

        if self.levaCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(11, False)

        if not self.holidaysCheckBox.isChecked():
            self.paymentsTableView.setColumnHidden(11, True)

        self.checkAllstate()

    def onPositionStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(3, not state)
        self.checkAllstate()

    def onPlaceStateChanged(self, state):
        self.paymentsTableView.setColumnHidden(4, not state)
        self.checkAllstate()

    def showCalendarDialog(self):
        calendarDialog = CustomCalendarDialog()
        Utils.calculatingIdealDialogShowPos(self, calendarDialog)
        if self.sender() == self.fromCalendarBtn:
            calendarDialog.calendarCustomWidget.setSelectedDate(self.fromDateEdit.date())
            calendarDialog.dateSelected.connect(self.updateFromDateLabel)
        elif self.sender() == self.toCalendarBtn:
            calendarDialog.calendarCustomWidget.setSelectedDate(self.fromDateEdit.date())
            calendarDialog.dateSelected.connect(self.updateToDateLabel)

        calendarDialog.exec_()

    def updateFromDateLabel(self, selectedDate):
        self.fromDateEdit.setDate(selectedDate)

    def updateToDateLabel(self, selectedDate):
        self.toDateEdit.setDate(selectedDate)

    def checkFromDate(self):
        if self.fromDateEdit.date() > self.toDateEdit.date():
            self.toDateEdit.setDate(self.fromDateEdit.date())

    def checkToDate(self):
        if self.toDateEdit.date() < self.fromDateEdit.date():
            self.fromDateEdit.setDate(self.toDateEdit.date())

    def showPaymentDetails(self, selectedRows):
        numberSelectedRows = len(selectedRows['payInEuro'])
        totalPayInLeva = 0
        totalPayInEuro = 0
        # avrLeva = 0
        # avrEuro = 0

        for key, row in selectedRows.items():
            if key == 'payInEuro':
                for item in row:
                    totalPayInEuro += float(item.data())
            else:
                for item in row:
                    totalPayInLeva += float(item.data())

        avrLeva = round(totalPayInLeva / len(selectedRows['payInLeva']), 2)
        avrEuro = round(totalPayInEuro / len(selectedRows['payInEuro']), 2)

        self.totalSelectedRows.setText(str(numberSelectedRows))
        self.totalLevaLabel.setText(str(round(totalPayInLeva, 2)))
        self.totalEuroLabel.setText(str(round(totalPayInEuro, 2)))
        self.avrLevaLabel.setText(str(avrLeva))
        self.avrEuroLabel.setText(str(avrEuro))

    def resetSelectedInfo(self):
        self.totalSelectedRows.setText("0")
        self.totalLevaLabel.setText("0.0")
        self.totalEuroLabel.setText("0.0")
        self.avrLevaLabel.setText("0.0")
        self.avrEuroLabel.setText("0.0")

    def setColumnWidths(self):
        self.paymentsTableView.setColumnWidth(0, 20)
        self.paymentsTableView.setColumnWidth(1, 40)
        self.paymentsTableView.setColumnWidth(2, 180)
        self.paymentsTableView.setColumnWidth(3, 100)
        self.paymentsTableView.setColumnWidth(12, 60)

    def refreshPaymentsTable(self, startDate=None, endDate=None):
        if not startDate:
            startDate = Utils.convertQDateToDate(self.fromDateEdit.date())
        else:
            startDate = Utils.convertQDateToDate(startDate)
        if not endDate:
            endDate = Utils.convertQDateToDate(self.toDateEdit.date())
        else:
            endDate = Utils.convertQDateToDate(endDate)
        rows = WoS.getInfoForPayments(startDate, endDate)
        self.tablePaymentsModel.setRowCount(0)
        for row in rows:
            rowToAdd = [
                QStandardItem(row['count']),
                QStandardItem(row['workerId']),
                QStandardItem(row['workerName']),
                QStandardItem(row['workerPosition']),
                QStandardItem(row['workerPlace']),
                QStandardItem(row['workingDays']),
                QStandardItem(row['totalShiftsTime']),
                QStandardItem(row['totalWeekEndWorkingDays']),
                QStandardItem(row['totalWeekendPayInLeva']),
                QStandardItem(row['totalWeekendPayInEuro']),
                QStandardItem(row['totalHolidaysWorking']),
                QStandardItem(row['totalHolidaysPayInLeva']),
                QStandardItem(row['totalHolidaysPayInEuro']),
                QStandardItem(row['totalPieces']),
                QStandardItem(row['displayOperTime']),
                QStandardItem(row['displayHourlyTime']),
                QStandardItem(row['displayOverTime']),
                QStandardItem(row['overtimePayInLeva']),
                QStandardItem(row['overtimePayInEuro']),
                QStandardItem(row['displayNightTime']),
                QStandardItem(row['nightPayInLeva']),
                QStandardItem(row['nightPayInEuro']),
                QStandardItem(row['efficiency']),
                QStandardItem(row['totalOvertimeForAllMins']),
                QStandardItem(row['totalOvertimeForAllMinsLeva']),
                QStandardItem(row['totalOvertimeForAllMinsEuro']),
                QStandardItem(row['payInLeva']),
                QStandardItem(row['payInEuro']),
                QStandardItem(row['totalPayInLeva']),
                QStandardItem(row['totalPayInEuro'])
            ]

            self.tablePaymentsModel.appendRow(rowToAdd)
        self.totalViewRows.setText(str(len(rows)))

    def closeEvent(self, event):
        self.mainWindow.closeAllPaymentsDetails()

    def openPaymentDetails(self, index):
        workerId = self.proxyModelPaymentsTable.mapToSource(index).siblingAtColumn(1).data(Qt.ItemDataRole.DisplayRole)
        totalLeva = self.proxyModelPaymentsTable.mapToSource(index).siblingAtColumn(28).data(Qt.ItemDataRole.DisplayRole)
        totalEuro = self.proxyModelPaymentsTable.mapToSource(index).siblingAtColumn(29).data(Qt.ItemDataRole.DisplayRole)
        startDate = self.fromDateEdit.date()
        endDate = self.toDateEdit.date()
        self.mainWindow.setPaymentsDetailsPage(workerId, startDate, endDate, totalLeva, totalEuro)

    def setProxyModel(self, proxyModel, model, table):
        proxyModel.setSourceModel(model)
        filterableHeaderView = FilterableHeaderView(Qt.Orientation.Horizontal, table)
        table.setHorizontalHeader(filterableHeaderView)
        filterableHeaderView.filterRequested.connect(partial(self.updateFilter, proxyModel, filterableHeaderView))
        table.setModel(proxyModel)
        table.setSortingEnabled(True)

    def updateFilter(self, proxyModel, header, column):
        if column < 5:
            # Create dictionaries to store column values and checked states if they don't exist
            if not hasattr(self, 'columnValues'):
                self.columnValues = {}
            if not hasattr(self, 'checkedStates'):
                self.checkedStates = {}
            if column not in self.checkBoxFiltering:
                self.checkBoxFiltering[column] = []

            items = []
            self.columnValues[column] = []

            for row in range(proxyModel.rowCount()):
                index = proxyModel.index(row, column)
                value = proxyModel.data(index, Qt.ItemDataRole.DisplayRole)
                if value not in self.columnValues[column]:
                    self.columnValues[column].append(value)
                    if value not in self.initialCheckBoxes:
                        self.initialCheckBoxes[value] = column

            for value in self.columnValues[column]:
                items.append(value)

            if items:
                menu = CustomSortingMenuWidget(header)
                for item in items:
                    checked = item in self.checkBoxFiltering[column]
                    menu.addItem(item, checked)
                menu.setMenuSize(len(items))
                menu.move(self.cursor().pos())
                menu.checkedCheckbox.connect(partial(self.applyFilter, proxyModel))

    def applyFilter(self, proxyModel, item):
        col = self.initialCheckBoxes[item.text()]
        if item.isChecked() and item.text() not in self.checkBoxFiltering[col]:
            self.checkBoxFiltering[col].append(item.text())
        elif not item.isChecked() and item.text() in self.checkBoxFiltering[col]:
            self.checkBoxFiltering[col].remove(item.text())

        if not self.checkBoxFiltering[col]:
            if col in proxyModel.columnFilters:
                del proxyModel.columnFilters[col]
        else:
            proxyModel.setFilterForColumn(col, self.checkBoxFiltering)
        proxyModel.invalidateFilter()
        self.totalViewRows.setText(str(proxyModel.rowCount()))

    def logout(self):
        self.logoutSignal.emit(True)
    