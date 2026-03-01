from PySide6.QtCore import Signal, QModelIndex
from PySide6.QtGui import QStandardItemModel, QStandardItem

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from PySide6.QtWidgets import QFileDialog

from app.ui.widgets.ui_customPaymentsDetailsWidget import *
from app.services.workerServices import WorkerServices as WoS
from app.models.customTreeView import CustomTreeView
from app.models.sortingModel import CaseInsensitiveProxyModel
from app.utils.utils import Utils


class CustomPaymentsDetailsWidget(QWidget, Ui_customPaymentsDetailsWidget):
    logoutSignal = Signal(bool)

    def __init__(self, mainWindow, workerId, startDate, endDate, totalPayInLeva, totalPayInEuro, user, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose, True)
        self.setWindowTitle(f'Детайли за работник с ID: {workerId}')
        self.mainWindow = mainWindow
        self.usernameLabel.setText(user)
        self.user = user

        self.workerId = workerId
        self.startDate = startDate
        self.endDate = endDate
        self.workerDetalsTreeView = CustomTreeView(self)
        self.totalPayInLeva = totalPayInLeva
        self.totalPayInEuro = totalPayInEuro

        self.rowIndx = 0
        self.excelParrentRows = []

        self.fromDateLineEdit.setText(self.startDate.toString('dd.MM.yyyy'))
        self.toDateLineEdit.setText(self.endDate.toString('dd.MM.yyyy'))

        self.setupWindowInfo()

        self.paymentsDetailsTableHolder.layout().addWidget(self.workerDetalsTreeView)
        self.tablePaymentDetailsModel = QStandardItemModel()
        self.tablePaymentDetailsNames = ['ID', 'Дата', 'Смяна', 'Прис. Вр.', 'Поръчки', 'Опер.', 'Вр. мин.',
                                         'Поч. дни', 'Поч. дни Лв.', 'Поч. дни €', 'В Празници', 'В Празници Лв.',
                                         'В Празници €', 'Почас.', 'Извънр.', 'Извънр. Лв.', 'Извънр. €', 'Нощен',
                                         'Нощен Лв.', 'Нощен €', 'Бр.', 'Ефект.', 'Извънр. Тотал', 'Извънр. Тотал Лв.',
                                         'Извънр. Тотал €', 'Зар. Лв.', 'Зар. €', 'Тотал лв.', 'Тотал €']

        for i, tableHeaderName in enumerate(self.tablePaymentDetailsNames):
            self.tablePaymentDetailsModel.setHorizontalHeaderItem(i, QStandardItem(tableHeaderName))
            self.tablePaymentDetailsModel.horizontalHeaderItem(i).setTextAlignment(Qt.AlignmentFlag.AlignLeft)
            self.tablePaymentDetailsModel.horizontalHeaderItem(i).setTextAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.proxyModelPaymentsDetailsTree = CaseInsensitiveProxyModel(numericColumns=[0, 3, 5, 6, 7, 8, 9, 10, 11, 12,
                                                                                       13, 14, 15, 16, 17, 18, 19, 20,
                                                                                       21, 22, 23, 24, 25, 26, 27, 28],
                                                                       dateColumns=[1],
                                                                       parent=self)

        self.setProxyModel(self.proxyModelPaymentsDetailsTree, self.tablePaymentDetailsModel, self.workerDetalsTreeView)
        # self.timePapersForDayTableView.setModel(self.tablePaymentsModel)
        self.workerDetalsTreeView.header().setDefaultSectionSize(70)
        self.workerDetalsTreeView.header().setStretchLastSection(True)
        self.setColumnWidths()
        self.refreshPaymentsDetailsTreeView()
        self.setInitialColumns()

        self.workerDetalsTreeView.selectedRows.connect(self.showPaymentDetails)
        self.workerDetalsTreeView.clearCurrentSelection.connect(self.resetSelectedInfo)

        self.levaCheckBox.stateChanged.connect(self.onLevaStateChanged)
        self.hourlyCheckBox.stateChanged.connect(self.onHourlyStateChanged)
        self.overtimeCheckBox.stateChanged.connect(self.onOvertimeStateChanged)
        self.nightTimeCheckBox.stateChanged.connect(self.onNightTimeStateChanged)
        self.selectAllCheckBox.stateChanged.connect(self.onSelectAllStateChanged)
        self.weekendCheckBox.stateChanged.connect(self.onWeekendStateChanged)
        self.holidaysCheckBox.stateChanged.connect(self.onHolidaysStateChanged)
        self.excelBtn.clicked.connect(self.exportToExcel)

        self.closeBtn.clicked.connect(self.close)
        self.logoutBtn.clicked.connect(self.logout)

    def exportToExcel(self):
        """Export the current table view data to Excel file"""
        workerName = self.workerNameLabel.text()
        workerLastName = self.workerLastNameLabel.text()
        workerNum = int(self.workerNumberLabel.text())

        # Get save file location from user
        fileName, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            os.path.join(os.path.expanduser("~"), "Documents",
                         f"PaymentDetails_{workerNum}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            "Excel Files (*.xlsx)"
        )

        if not fileName:
            return  # User canceled

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Details Report"

        # Add report title and date range
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = (f"Payment {workerNum}: {workerName} {workerLastName} Report " +
                            f"({self.fromDateLineEdit.text()} - " +
                            f"{self.toDateLineEdit.text()})")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        # Add headers
        header_fill = PatternFill(start_color="9FABB3", end_color="9FABB3", fill_type="solid")

        columnHeaders = self.tablePaymentDetailsNames.copy()

        if not self.holidaysCheckBox.isChecked():
            columnHeaders.remove('В Празници')
            columnHeaders.remove('В Празници €')
            columnHeaders.remove('В Празници Лв.')

        if self.holidaysCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('В Празници Лв.')

        if not self.weekendCheckBox.isChecked():
            columnHeaders.remove('Поч. дни')
            columnHeaders.remove('Поч. дни Лв.')
            columnHeaders.remove('Поч. дни €')

        if self.weekendCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('Поч. дни Лв.')

        if not self.overtimeCheckBox.isChecked():
            columnHeaders.remove('Извънр.')
            columnHeaders.remove('Извънр. Лв.')
            columnHeaders.remove('Извънр. €')

        if self.overtimeCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('Извънр. Лв.')

        if not self.nightTimeCheckBox.isChecked():
            columnHeaders.remove('Нощен')
            columnHeaders.remove('Нощен Лв.')
            columnHeaders.remove('Нощен €')

        if self.nightTimeCheckBox.isChecked() and not self.levaCheckBox.isChecked():
            columnHeaders.remove('Нощен Лв.')

        if not self.hourlyCheckBox.isChecked():
            columnHeaders.remove('Почас.')

        if not self.levaCheckBox.isChecked():
            columnHeaders.remove('Зар. Лв.')
            columnHeaders.remove('Тотал лв.')
            columnHeaders.remove('Извънр. Тотал Лв.')

        # if not self.weekendHolidaysCheckBox.isChecked():
        #     columnHeaders.remove('Поч./Празн. дни')
        #
        # if not self.overtimeCheckBox.isChecked():
        #     columnHeaders.remove('Изв. Раб.')
        #
        # if not self.nightTimeCheckBox.isChecked():
        #     columnHeaders.remove('Нощен труд')
        #
        # if not self.hourlyCheckBox.isChecked():
        #     columnHeaders.remove('Поч. Раб.')

        for col_idx, header_text in enumerate(columnHeaders):
            col_letter = get_column_letter(col_idx + 1)
            cell = ws[f"{col_letter}3"]
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        # self.excelTableNames = self.tablePaymentsNames

        # Add data rows
        row_idx = 4  # Start from row 4 (after headers)
        self.rowIndx = row_idx
        self.excelParrentRows = []
        self.writeTreeRowsToSheet(ws, self.proxyModelPaymentsDetailsTree, QModelIndex())
        row_idx = self.rowIndx
        # Add summary at the bottom
        summary_row = row_idx + 1
        ws.cell(row=summary_row, column=1, value="Total:").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=self.totalViewRows.text()).font = Font(bold=True)

        # Add SUM formulas for Leva (column 13) and Euro (column 14)
        leva_col = get_column_letter(len(columnHeaders) - 1)  # Column M in Excel (index 12 + 1)
        euro_col = get_column_letter(len(columnHeaders))  # Column N in Excel (index 13 + 1)

        totalLeva = "=("
        totalEuro = "=("
        for i in range(len(self.excelParrentRows)):
            if i == len(self.excelParrentRows) - 1:
                totalLeva += f"{leva_col}{self.excelParrentRows[i]})"
                totalEuro += f"{euro_col}{self.excelParrentRows[i]})"
            else:
                totalLeva += f"{leva_col}{self.excelParrentRows[i]}+"
                totalEuro += f"{euro_col}{self.excelParrentRows[i]}+"

        # Create SUM formulas that reference the data range
        leva_formula = totalLeva
        euro_formula = totalEuro

        # Add the formulas to the summary row
        leva_cell = ws.cell(row=summary_row, column=len(columnHeaders) - 1, value=leva_formula)
        euro_cell = ws.cell(row=summary_row, column=len(columnHeaders), value=euro_formula)

        # Apply bold formatting to the summary cells
        leva_cell.font = Font(bold=True)
        euro_cell.font = Font(bold=True)

        # Save the workbook
        try:
            wb.save(fileName)
            from app.ui.messagesManager import MessageManager as MM
            MM.showOnWidget(self, f"Успешен запис на {fileName}", "success")
        except Exception as e:
            from app.ui.messagesManager import MessageManager as MM
            MM.showOnWidget(self, f"Грешка при записване: {str(e)}", "error")

    def writeTreeRowsToSheet(self, ws, model, parentIndex, isChild=False):
        cell_fill = PatternFill(start_color="C1C4C9", end_color="C1C4C9", fill_type="solid")

        for row in range(model.rowCount(parentIndex)):
            viewColumns = 0
            for col in range(model.columnCount(parentIndex)):
                if self.workerDetalsTreeView.isColumnHidden(col):
                    continue

                # Get data from the proxy model
                index = model.index(row, col, parentIndex)
                value = model.data(index, Qt.ItemDataRole.DisplayRole)

                # Convert to appropriate type
                if col in [0, 5, 20, 21]:  # Integer columns
                    try:
                        value = int(value) if value else 0
                    except (ValueError, TypeError):
                        pass
                elif col in [3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15,
                             16, 17, 18, 19, 22, 23, 24, 25, 26, 27, 28]:  # Float columns
                    try:
                        value = float(value) if value else 0.0
                    except (ValueError, TypeError):
                        pass

                cell = ws.cell(row=self.rowIndx, column=viewColumns + 1, value=value)
                if not isChild and self.includeSubrowCheckBox.isChecked():
                    cell.fill = cell_fill
                viewColumns += 1
            if not isChild:
                self.excelParrentRows.append(self.rowIndx)
            self.rowIndx += 1
            if model.hasChildren(model.index(row, 0, parentIndex)) and self.includeSubrowCheckBox.isChecked():
                self.writeTreeRowsToSheet(ws, model, model.index(row, 0, parentIndex), True)

    def showPaymentDetails(self, selectedRows):
        numberSelectedRows = len(selectedRows['payInEuro'])
        totalPayInLeva = 0
        totalPayInEuro = 0
        avrLeva = 0
        avrEuro = 0

        for key, row in selectedRows.items():
            if key == 'payInEuro':
                for item in row:
                    totalPayInEuro += float(item.data())
            else:
                for item in row:
                    totalPayInLeva += float(item.data())

        if len(selectedRows['payInLeva']) > 0:
            avrLeva = round(totalPayInLeva / len(selectedRows['payInLeva']), 2)
            avrEuro = round(totalPayInEuro / len(selectedRows['payInEuro']), 2)

        self.totalSelectedRows.setText(str(numberSelectedRows))
        self.totalLevaLabel.setText(str(round(totalPayInLeva, 2)))
        self.totalEuroLabel.setText(str(round(totalPayInEuro, 2)))
        self.avrLevaLabel.setText(str(avrLeva))
        self.avrEuroLabel.setText(str(avrEuro))

    def resetSelectedInfo(self):
        self.totalSelectedRows.setText("0")
        self.totalLevaLabel.setText("0.0")
        self.totalEuroLabel.setText("0.0")
        self.avrLevaLabel.setText("0.0")
        self.avrEuroLabel.setText("0.0")

    def setColumnWidths(self):
        self.workerDetalsTreeView.setColumnWidth(1, 85)
        self.workerDetalsTreeView.setColumnWidth(3, 85)
        # self.workerDetalsTreeView.setColumnWidth(4, 52)
        # self.workerDetalsTreeView.setColumnWidth(5, 68)
        # self.workerDetalsTreeView.setColumnWidth(6, 85)
        # self.workerDetalsTreeView.setColumnWidth(7, 72)
        # self.workerDetalsTreeView.setColumnWidth(8, 72)
        # self.workerDetalsTreeView.setColumnWidth(9, 42)
        # self.workerDetalsTreeView.setColumnWidth(10, 60)

    # ['ID'0, 'Дата'1, 'Смяна'2, 'Прис. Вр.'3, 'Поръчки'4, 'Опер.'5, 'Вр. мин.'6,
    #  'Поч. дни'7, 'Поч. дни Лв.'8, 'Поч. дни €'9, 'В Празници'10, 'В Празници Лв.'11,
    #  'В Празници €'12, 'Почас.'13, 'Извънр.'14, 'Извънр. Лв.'15, 'Извънр. €'16, 'Нощен'17,
    #  'Нощен Лв.'18, 'Нощен €'19, 'Бр.'20, 'Ефект.'21, 'Извънр. Тотал'22, 'Извънр. Тотал Лв.'23,
    #  'Извънр. Тотал €'24, 'Зар. Лв.'25, 'Зар. €'26, 'Тотал лв.'27, 'Тотал €'28]

    def setInitialColumns(self):
        for i in range(len(self.tablePaymentDetailsNames)):
            if i in [0, 1, 2, 3, 4, 5, 6, 20, 21, 22, 24, 26, 28]:
                pass
            else:
                self.workerDetalsTreeView.setColumnHidden(i, True)

    def checkAllstate(self):
        if (self.hourlyCheckBox.isChecked() and
                self.overtimeCheckBox.isChecked() and
                self.nightTimeCheckBox.isChecked() and
                self.weekendCheckBox.isChecked() and
                self.holidaysCheckBox.isChecked() and
                self.levaCheckBox.isChecked()):
            self.selectAllCheckBox.setChecked(True)
        else:
            self.selectAllCheckBox.blockSignals(True)
            self.selectAllCheckBox.setChecked(False)
            self.selectAllCheckBox.blockSignals(False)

    def onSelectAllStateChanged(self, state):
        self.hourlyCheckBox.setChecked(state)
        self.overtimeCheckBox.setChecked(state)
        self.nightTimeCheckBox.setChecked(state)
        self.weekendCheckBox.setChecked(state)
        self.holidaysCheckBox.setChecked(state)
        self.levaCheckBox.setChecked(state)

    def onLevaStateChanged(self, state):
        self.workerDetalsTreeView.setColumnHidden(23, not state)
        self.workerDetalsTreeView.setColumnHidden(25, not state)
        self.workerDetalsTreeView.setColumnHidden(27, not state)

        if self.weekendCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(8, not state)

        if self.holidaysCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(11, not state)

        if self.overtimeCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(15, not state)

        if self.nightTimeCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(18, not state)

        self.checkAllstate()

    def onHolidaysStateChanged(self, state):
        self.workerDetalsTreeView.setColumnHidden(10, not state)
        self.workerDetalsTreeView.setColumnHidden(12, not state)

        if self.levaCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(11, False)

        if not self.holidaysCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(11, True)

    def onWeekendStateChanged(self, state):
        self.workerDetalsTreeView.setColumnHidden(7, not state)
        self.workerDetalsTreeView.setColumnHidden(9, not state)

        if self.levaCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(8, False)

        if not self.weekendCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(8, True)

        self.checkAllstate()

    def onHourlyStateChanged(self, state):
        self.workerDetalsTreeView.setColumnHidden(13, not state)
        self.checkAllstate()

    def onOvertimeStateChanged(self, state):
        self.workerDetalsTreeView.setColumnHidden(14, not state)
        self.workerDetalsTreeView.setColumnHidden(16, not state)

        if self.levaCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(15, False)

        if not self.overtimeCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(15, True)

        self.checkAllstate()

    def onNightTimeStateChanged(self, state):
        self.workerDetalsTreeView.setColumnHidden(17, not state)
        self.workerDetalsTreeView.setColumnHidden(19, not state)

        if self.levaCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(18, False)

        if not self.nightTimeCheckBox.isChecked():
            self.workerDetalsTreeView.setColumnHidden(18, True)

        self.checkAllstate()

    def refreshPaymentsDetailsTreeView(self):
        self.tablePaymentDetailsModel.setRowCount(0)
        startDate = QDate.toString(self.startDate, 'dd.MM.yyyy')
        endDate = QDate.toString(self.endDate, 'dd.MM.yyyy')
        paymentsData = WoS.getPaymentsDetailsForWorker(self.workerId, startDate, endDate)
        if paymentsData:
            totalRows = 0

            self.levaPerMinDayLabel.setText(str(paymentsData[0]['paymentInLv']))
            self.euroPerMinDayLabel.setText(str(paymentsData[0]['paymentInEuro']))
            self.levaPerMinNightLabel.setText(str(paymentsData[0]['nightPayInLeva']))
            self.euroPerMinNightLabel.setText(str(paymentsData[0]['nightPayInEuro']))
            self.paymentLevaLabel.setText(f'{self.totalPayInLeva} лв')
            self.paymentEuroLabel.setText(f'{self.totalPayInEuro} €')

            parentRow = None
            currentParent = 0
            for row in paymentsData:
                if row['row'].split('_')[0] == 'parent':
                    currentParent = row['row'].split('_')[1]
                    parrentRowToAdd = QStandardItem(str(row['payment']))
                    parentRow = parrentRowToAdd

                    # ['ID'0, 'Дата'1, 'Смяна'2, 'Прис. Вр.'3, 'Поръчки'4, 'Опер.'5, 'Вр. мин.'6,
                    #  'Поч. дни'7, 'Поч. дни Лв.'8, 'Поч. дни €'9, 'В Празници'10, 'В Празници Лв.'11,
                    #  'В Празници €'12, 'Почас.'13, 'Извънр.'14, 'Извънр. Лв.'15, 'Извънр. €'16, 'Нощен'17,
                    #  'Нощен Лв.'18, 'Нощен €'19, 'Бр.'20, 'Ефект.'21, 'Извънр. Тотал'22, 'Извънр. Тотал Лв.'23,
                    #  'Извънр. Тотал €'24, 'Зар. Лв.'25, 'Зар. €'26, 'Тотал лв.'27, 'Тотал €'28]

                    # (row, payment, date, shift, shiftTime, ordersCount, operationsCount, totalTime, weekends,
                    #  weekendsLeva, weekendsEuro, holidays, holidaysLeva, hildaysEuro, totalHourlyTime,
                    #  totalOvertime,
                    #  overtimeLeva, overtimeEuro, totalNightTime, nightTimeLeva, nightTimeEuro, totalPieces,
                    #  efficency, totalOvertimeMins, totalOvertimeMinsLeva, totalOvertimeMinsEuro,
                    #  workingLeva, workingEuro, totalPaymentInLev, totalPaymentInEuro):

                    parrentRowForAppend = [
                        parrentRowToAdd,
                        QStandardItem(row['date']),
                        QStandardItem(row['shift']),
                        QStandardItem(row['shiftTime']),
                        QStandardItem(row['ordersCount']),
                        QStandardItem(row['operationsCount']),
                        QStandardItem(row['totalTime']),
                        QStandardItem(row['weekends']),
                        QStandardItem(row['weekendsLeva']),
                        QStandardItem(row['weekendsEuro']),
                        QStandardItem(row['holidays']),
                        QStandardItem(row['holidaysLeva']),
                        QStandardItem(row['holidaysEuro']),
                        QStandardItem(row['totalHourlyTime']),
                        QStandardItem(row['totalOvertime']),
                        QStandardItem(row['overtimeLeva']),
                        QStandardItem(row['overtimeEuro']),
                        QStandardItem(row['totalNightTime']),
                        QStandardItem(row['nightTimeLeva']),
                        QStandardItem(row['nightTimeEuro']),
                        QStandardItem(row['totalPieces']),
                        QStandardItem(row['efficency']),
                        QStandardItem(row['totalOvertimeMins']),
                        QStandardItem(row['totalOvertimeMinsLeva']),
                        QStandardItem(row['totalOvertimeMinsEuro']),
                        QStandardItem(row['workingLeva']),
                        QStandardItem(row['workingEuro']),
                        QStandardItem(row['totalPaymentInLev']),
                        QStandardItem(row['totalPaymentInEuro'])
                    ]
                    totalRows += 1
                    self.tablePaymentDetailsModel.appendRow(parrentRowForAppend)

                elif row['row'] == f'child_{currentParent}':
                    childRowForAppend = [
                        QStandardItem(str(row['payment'])),
                        QStandardItem(row['date']),
                        QStandardItem(row['shift']),
                        QStandardItem(row['shiftTime']),
                        QStandardItem(row['ordersCount']),
                        QStandardItem(row['operationsCount']),
                        QStandardItem(row['totalTime']),
                        QStandardItem(row['weekends']),
                        QStandardItem(row['weekendsLeva']),
                        QStandardItem(row['weekendsEuro']),
                        QStandardItem(row['holidays']),
                        QStandardItem(row['holidaysLeva']),
                        QStandardItem(row['holidaysEuro']),
                        QStandardItem(row['totalHourlyTime']),
                        QStandardItem(row['totalOvertime']),
                        QStandardItem(row['overtimeLeva']),
                        QStandardItem(row['overtimeEuro']),
                        QStandardItem(row['totalNightTime']),
                        QStandardItem(row['nightTimeLeva']),
                        QStandardItem(row['nightTimeEuro']),
                        QStandardItem(row['totalPieces']),
                        QStandardItem(row['efficency']),
                        QStandardItem(row['totalOvertimeMins']),
                        QStandardItem(row['totalOvertimeMinsLeva']),
                        QStandardItem(row['totalOvertimeMinsEuro']),
                        QStandardItem(row['workingLeva']),
                        QStandardItem(row['workingEuro']),
                        QStandardItem(row['totalPaymentInLev']),
                        QStandardItem(row['totalPaymentInEuro'])
                    ]
                    parentRow.appendRow(childRowForAppend)

            self.totalViewRows.setText(str(totalRows))
            self.proxyModelPaymentsDetailsTree.sort(1, Qt.SortOrder.DescendingOrder)

    def setProxyModel(self, proxyModel, model, table):
        proxyModel.setSourceModel(model)
        table.setModel(proxyModel)
        table.setSortingEnabled(True)

    def setupWindowInfo(self):
        workerData = WoS.getWorkerDataForPayments(self.workerId)
        self.workerNameLabel.setText(workerData['firstName'])
        self.workerLastNameLabel.setText(workerData['lastName'])
        self.workerNumberLabel.setText(str(workerData['id']))
        self.workerPlaceLabel.setText(workerData['place'])
        self.workerPositionLabel.setText(workerData['position'])

    def logout(self):
        self.logoutSignal.emit(True)
